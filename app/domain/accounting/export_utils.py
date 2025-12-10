"""Export utilities for financial reports."""
from typing import Dict
from datetime import date
from decimal import Decimal
import os
from io import BytesIO

# PDF generation
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# Excel generation
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


class ReportExporter:
    """Utility class for exporting financial reports to PDF and Excel."""
    
    @staticmethod
    def export_balance_sheet_to_pdf(balance_sheet: Dict, output_path: str) -> str:
        """Export balance sheet to PDF."""
        doc = SimpleDocTemplate(
            output_path,
            pagesize=A4,
            rightMargin=2*cm,
            leftMargin=2*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )
        
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#667eea'),
            spaceAfter=30,
            alignment=1
        )
        elements.append(Paragraph("BALANÇ DE SITUACIÓ", title_style))
        
        date_text = f"Data de tancament: {balance_sheet['end_date'].strftime('%d/%m/%Y')}" if balance_sheet.get('end_date') else "Situació actual"
        elements.append(Paragraph(date_text, styles['Normal']))
        elements.append(Spacer(1, 20))
        
        # Helper to generate section table
        def add_section_table(title, data_dict, total_amount):
            elements.append(Paragraph(title, styles['Heading3']))
            
            table_data = []
            # Header
            table_data.append(['Compte', 'Nom', 'Import (€)'])
            
            # Groups
            for group_name, group_data in data_dict['groups'].items():
                table_data.append([group_name, '', '']) # Group Header
                for account in group_data['accounts']:
                    table_data.append([
                        account['code'],
                        account['name'],
                        f"{float(account['balance']):.2f}"
                    ])
                table_data.append(['', f"Total {group_name}", f"{float(group_data['total']):.2f}"])
            
            table_data.append(['', f"TOTAL {title.upper()}", f"{float(total_amount):.2f}"])
            
            t = Table(table_data, colWidths=[3*cm, 10*cm, 4*cm])
            
            # Dynamic styling
            style = [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ]
            
            row_idx = 1
            for group_name, group_data in data_dict['groups'].items():
                # Group Header Style
                style.append(('BACKGROUND', (0, row_idx), (-1, row_idx), colors.lightgrey))
                style.append(('FONTNAME', (0, row_idx), (-1, row_idx), 'Helvetica-Bold'))
                row_idx += 1
                for _ in group_data['accounts']:
                    row_idx += 1
                # Group Total Style
                style.append(('FONTNAME', (0, row_idx), (-1, row_idx), 'Helvetica-Oblique'))
                row_idx += 1
            
            # Grand Total Style
            style.append(('BACKGROUND', (0, row_idx), (-1, row_idx), colors.HexColor('#d4d8ff')))
            style.append(('FONTNAME', (0, row_idx), (-1, row_idx), 'Helvetica-Bold'))
            
            t.setStyle(TableStyle(style))
            elements.append(t)
            elements.append(Spacer(1, 15))

        # ACTIU
        elements.append(Paragraph("A) ACTIU", styles['Heading2']))
        add_section_table("Actiu No Corrent", balance_sheet['actiu']['no_corrent'], balance_sheet['actiu']['no_corrent']['total'])
        add_section_table("Actiu Corrent", balance_sheet['actiu']['corrent'], balance_sheet['actiu']['corrent']['total'])
        
        # Total Actiu
        elements.append(Paragraph(f"TOTAL ACTIU: {float(balance_sheet['actiu']['total']):.2f} €", styles['Heading3']))
        elements.append(Spacer(1, 20))
        
        # PASSIU
        elements.append(Paragraph("B) PATRIMONI NET I PASSIU", styles['Heading2']))
        add_section_table("Patrimoni Net", balance_sheet['patrimoni_net_i_passiu']['patrimoni_net'], balance_sheet['patrimoni_net_i_passiu']['patrimoni_net']['total'])
        add_section_table("Passiu No Corrent", balance_sheet['patrimoni_net_i_passiu']['passiu_no_corrent'], balance_sheet['patrimoni_net_i_passiu']['passiu_no_corrent']['total'])
        add_section_table("Passiu Corrent", balance_sheet['patrimoni_net_i_passiu']['passiu_corrent'], balance_sheet['patrimoni_net_i_passiu']['passiu_corrent']['total'])

        elements.append(Paragraph(f"TOTAL PATRIMONI NET I PASSIU: {float(balance_sheet['patrimoni_net_i_passiu']['total']):.2f} €", styles['Heading3']))
        
        doc.build(elements)
        return output_path
    
    @staticmethod
    def export_balance_sheet_to_excel(balance_sheet: Dict, output_path: str) -> str:
        """Export balance sheet to Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Balanç de Situació"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        title_font = Font(bold=True, size=16, color="667eea")
        total_font = Font(bold=True, size=11)
        total_fill = PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        row = 1
        
        # Title
        ws.merge_cells(f'A{row}:C{row}')
        cell = ws[f'A{row}']
        cell.value = "BALANÇ DE SITUACIÓ"
        cell.font = title_font
        cell.alignment = Alignment(horizontal='center')
        row += 1
        
        # Date
        date_text = f"Data: {balance_sheet['end_date'].strftime('%d/%m/%Y')}" if balance_sheet.get('end_date') else "Situació actual"
        ws.merge_cells(f'A{row}:C{row}')
        ws[f'A{row}'].value = date_text
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        row += 2
        
        def add_excel_section(title, data_dict):
            nonlocal row
            ws[f'A{row}'].value = title
            ws[f'A{row}'].font = Font(bold=True, size=14)
            row += 1
            
            # Header
            headers = ["Compte", "Nom", "Import (€)"]
            for col_idx, text in enumerate(headers, 1):
                c = ws.cell(row=row, column=col_idx, value=text)
                c.font = header_font
                c.fill = header_fill
                c.border = border
            row += 1
            
            # Groups
            for group_name, group_data in data_dict['groups'].items():
                # Group Header
                ws.merge_cells(f'A{row}:C{row}')
                ws[f'A{row}'] = group_name
                ws[f'A{row}'].font = Font(bold=True, color="444444")
                ws[f'A{row}'].fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                row += 1
                
                for account in group_data['accounts']:
                    ws[f'A{row}'] = account['code']
                    ws[f'B{row}'] = account['name']
                    c = ws[f'C{row}']
                    c.value = float(account['balance'])
                    c.number_format = '#,##0.00'
                    for col in ['A', 'B', 'C']:
                        ws[f'{col}{row}'].border = border
                    row += 1
                    
                # Group Total
                ws[f'B{row}'] = f"Total {group_name}"
                c = ws[f'C{row}']
                c.value = float(group_data['total'])
                c.number_format = '#,##0.00'
                ws[f'C{row}'].font = Font(italic=True)
                row += 1

            # Section Total
            ws[f'B{row}'] = f"Total {title}"
            c = ws[f'C{row}']
            c.value = float(data_dict['total']) # This logic was wrong in previous code, needed data_dict['total']
            c.number_format = '#,##0.00'
            c.font = total_font
            c.fill = total_fill
            row += 2

        # ACTIU
        add_excel_section("Actiu No Corrent", balance_sheet['actiu']['no_corrent'])
        add_excel_section("Actiu Corrent", balance_sheet['actiu']['corrent'])
        
        # Total Actiu
        ws[f'B{row}'] = "TOTAL ACTIU"
        ws[f'C{row}'] = float(balance_sheet['actiu']['total'])
        ws[f'C{row}'].font = Font(bold=True, size=12)
        ws[f'C{row}'].fill = PatternFill(start_color="D4D8FF", end_color="D4D8FF", fill_type="solid")
        row += 2
        
        # PASSIU
        add_excel_section("Patrimoni Net", balance_sheet['patrimoni_net_i_passiu']['patrimoni_net'])
        add_excel_section("Passiu No Corrent", balance_sheet['patrimoni_net_i_passiu']['passiu_no_corrent'])
        add_excel_section("Passiu Corrent", balance_sheet['patrimoni_net_i_passiu']['passiu_corrent'])
        
        # Total Passiu
        ws[f'B{row}'] = "TOTAL PATRIMONI NET I PASSIU"
        ws[f'C{row}'] = float(balance_sheet['patrimoni_net_i_passiu']['total'])
        ws[f'C{row}'].font = Font(bold=True, size=12)
        ws[f'C{row}'].fill = PatternFill(start_color="D4D8FF", end_color="D4D8FF", fill_type="solid")
        
        # Adjust widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 15
        
        wb.save(output_path)
        return output_path
    
    @staticmethod
    def export_profit_loss_to_pdf(profit_loss: Dict, output_path: str) -> str:
        """Export profit & loss statement to PDF."""
        doc = SimpleDocTemplate(output_path, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#667eea'),
            spaceAfter=30,
            alignment=1
        )
        elements.append(Paragraph("COMPTE DE PÈRDUES I GUANYS", title_style))
        
        # Period
        if profit_loss.get('start_date') and profit_loss.get('end_date'):
            period_text = f"Període: {profit_loss['start_date'].strftime('%d/%m/%Y')} - {profit_loss['end_date'].strftime('%d/%m/%Y')}"
        else:
            period_text = "Tot el període"
        elements.append(Paragraph(period_text, styles['Normal']))
        elements.append(Spacer(1, 20))
        
        elements.append(Paragraph("A) OPERACIONS CONTINUADES", styles['Heading2']))
        
        table_data = [['Partida', 'Import (€)']]
        style = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]
        
        row_idx = 1
        
        # Sort groups
        sorted_groups = sorted(profit_loss['groups'].items())
        
        for group_name, group_data in sorted_groups:
            # Group Header
            table_data.append([group_name, f"{float(group_data['total']):.2f}"])
            style.append(('BACKGROUND', (0, row_idx), (-1, row_idx), colors.whitesmoke))
            style.append(('FONTNAME', (0, row_idx), (-1, row_idx), 'Helvetica-Bold'))
            row_idx += 1
            
            # Items
            for item in group_data['items']:
                table_data.append([f"   {item['code']} {item['name']}", f"{float(item['amount']):.2f}"])
                style.append(('TEXTCOLOR', (0, row_idx), (0, row_idx), colors.dimgrey))
                row_idx += 1
                
        # Results
        def add_result_row(label, amount, is_final=False):
            table_data.append([label, f"{float(amount):.2f}"])
            bg_color = colors.HexColor('#d4edda') if amount >= 0 else colors.HexColor('#f8d7da')
            if is_final:
                bg_color = colors.HexColor('#c3e6cb') if amount >= 0 else colors.HexColor('#f5c6cb')
                
            nonlocal row_idx # Use row_idx from outer scope
            style.append(('BACKGROUND', (0, row_idx), (-1, row_idx), bg_color))
            style.append(('FONTNAME', (0, row_idx), (-1, row_idx), 'Helvetica-Bold'))
            row_idx += 1

        add_result_row("A.1) RESULTAT D'EXPLOTACIÓ", profit_loss['resultat_explotacio'])
        add_result_row("A.2) RESULTAT FINANCER", profit_loss['resultat_financer'])
        add_result_row("A.3) RESULTAT ABANS D'IMPOSTOS", profit_loss['resultat_abans_impostos'])
        add_result_row("A.4) RESULTAT DE L'EXERCICI", profit_loss['resultat_exercici'], is_final=True)
        
        t = Table(table_data, colWidths=[12*cm, 4*cm])
        t.setStyle(TableStyle(style))
        elements.append(t)
        
        doc.build(elements)
        return output_path
    
    @staticmethod
    def export_profit_loss_to_excel(profit_loss: Dict, output_path: str) -> str:
        """Export profit & loss statement to Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Compte PyG"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        title_font = Font(bold=True, size=16, color="667eea")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        positive_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
        negative_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
        
        row = 1
        
        # Title
        ws.merge_cells(f'A{row}:C{row}')
        cell = ws[f'A{row}']
        cell.value = "COMPTE DE PÈRDUES I GUANYS"
        cell.font = title_font
        cell.alignment = Alignment(horizontal='center')
        row += 1
        
        # Period
        if profit_loss.get('start_date') and profit_loss.get('end_date'):
            period_text = f"Període: {profit_loss['start_date'].strftime('%d/%m/%Y')} - {profit_loss['end_date'].strftime('%d/%m/%Y')}"
        else:
            period_text = "Tot el període"
        ws.merge_cells(f'A{row}:C{row}')
        ws[f'A{row}'].value = period_text
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        row += 2
        
        ws[f'A{row}'].value = "A) OPERACIONS CONTINUADES"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        # Header
        ws[f'A{row}'] = "Partida"
        ws[f'B{row}'] = "Import (€)"
        for col in ['A', 'B']:
            ws[f'{col}{row}'].font = header_font
            ws[f'{col}{row}'].fill = header_fill
            ws[f'{col}{row}'].border = border
        
        row += 1
        
        # Sort groups
        sorted_groups = sorted(profit_loss['groups'].items())
        
        for group_name, group_data in sorted_groups:
            # Group Header
            ws[f'A{row}'] = group_name
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
            
            c = ws[f'B{row}']
            c.value = float(group_data['total'])
            c.number_format = '#,##0.00'
            c.font = Font(bold=True)
            c.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
            
            ws[f'A{row}'].border = border
            ws[f'B{row}'].border = border
            row += 1
            
            # Items
            for item in group_data['items']:
                ws[f'A{row}'] = f"   {item['code']} {item['name']}"
                ws[f'A{row}'].font = Font(color="555555")
                c = ws[f'B{row}']
                c.value = float(item['amount'])
                c.number_format = '#,##0.00'
                c.font = Font(color="555555")
                
                ws[f'A{row}'].border = border
                ws[f'B{row}'].border = border
                row += 1

        # Results Helper
        def add_result_row(label, amount, is_final=False):
            nonlocal row
            ws[f'A{row}'] = label
            c = ws[f'B{row}']
            c.value = float(amount)
            c.number_format = '#,##0.00'
            
            fill = positive_fill if amount >= 0 else negative_fill
            if is_final:
                fill = PatternFill(start_color="C3E6CB", end_color="C3E6CB", fill_type="solid") if amount >= 0 else PatternFill(start_color="F5C6CB", end_color="F5C6CB", fill_type="solid")
            
            ws[f'A{row}'].fill = fill
            ws[f'B{row}'].fill = fill
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'].font = Font(bold=True)
            ws[f'A{row}'].border = border
            ws[f'B{row}'].border = border
            
            row += 1

        add_result_row("A.1) RESULTAT D'EXPLOTACIÓ", profit_loss['resultat_explotacio'])
        add_result_row("A.2) RESULTAT FINANCER", profit_loss['resultat_financer'])
        add_result_row("A.3) RESULTAT ABANS D'IMPOSTOS", profit_loss['resultat_abans_impostos'])
        add_result_row("A.4) RESULTAT DE L'EXERCICI", profit_loss['resultat_exercici'], is_final=True)
        
        # Adjust widths
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 15
        
        wb.save(output_path)
        return output_path
